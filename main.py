# @Author: oliver
# @Date:   2021-03-16 10:27:59
# @Last Modified by:   oliver
# @Last Modified time: 2022-05-04 13:29
import os
from pandas.io.html import read_html
#region 模組導入
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd 
import datetime 
import openpyxl
from Logger import create_logger 
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import jiragetbug
from argparse import ArgumentParser

logger = create_logger(r"\AutoTest", 'main_log')

#endregion
delay=0.3
excel_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),r'空白範本.xlsx')

def csv_to_xlsx_pd(filepath):
    try:
        logger.info('開始將html轉成xlsx檔')
        csv = pd.read_html(os.path.join(filepath,r'tempfile123.html'),index_col=False,encoding='UTF-8')[0]
        csv.columns=['結果','用例編號','所屬模組','用例標題',"實際情況","測試人"]
        csv['結果']=csv['結果'].map({'通过':'通過','失败':'失敗','阻塞':'阻塞'})
        csv.to_excel(os.path.join(filepath,r'tempfile123.xlsx'), sheet_name='用例',index=False)
        
        logger.info('xlsx轉換成功')
        return os.path.join(filepath,r'tempfile123.xlsx')
    except Exception as e:
        logger.error(f"html轉成xlsx檔錯誤/n =========== /n {e} /n ===========", exc_info=True)


def generate_report(downloadpath,targetpath,BUGdf,jiraurl=''):
    logger.info('開始生成報表')
    try:
        wb1=openpyxl.load_workbook(excel_path)
        testlistpath=""
        wsbug=wb1.create_sheet("BUG表")
        ws1=wb1.create_sheet("用例")
        sheet=wb1['總結']
        sheet['B4']=jiraurl
        
        rows=dataframe_to_rows(BUGdf,index=False)
        try:
            logger.info('寫入BUG')
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    wsbug.cell(row=r_idx, column=c_idx, value=value)
            logger.info('BUG寫入完成')
        except:
            logger.warning('本次未寫入BUG', exc_info=True)
            pass

        try:
            logger.warning('寫入測試用例')
            testlistpath=csv_to_xlsx_pd(downloadpath)
            wb2=openpyxl.load_workbook(testlistpath)
            ws2=wb2['用例']
            for row in ws2:
                for cell in row:
                        ws1[cell.coordinate].value = cell.value
            logger.warning('測試用例寫入完成')
        except:
            logger.warning('本次未寫入測試用例', exc_info=True)
        
        
        wsbug.sheet_state='hidden'
        ws1.sheet_state='hidden'
        testcase=pd.read_excel(testlistpath)

        outputtext=f"本次共驗證【{testcase['結果'].value_counts().sum()}】個測試案例"
        try:
            outputtext+=f"，通過【{testcase['結果'].value_counts()['通過']}】個"
        except:
            pass
        try:
            outputtext+=f"，阻塞【{testcase['結果'].value_counts()['阻塞']}】個"
        except:
            pass
        try:
            outputtext+=f"，失敗【{testcase['結果'].value_counts()['失敗']}】個"
        except:
            pass
        outputtext+="""。
"""
        
        try:
            outputtext+=f"本次共發現【{BUGdf['BUG等級'].value_counts().sum()}】個BUG"
        except:
            outputtext+="本次未發現BUG"

        try:
            outputtext+=f"，A級BUG共【{BUGdf['BUG等級'].value_counts()['A']}】個"
        except:
            pass
        try:
            outputtext+=f"，B級BUG共【{BUGdf['BUG等級'].value_counts()['B']}】個"
        except:
            pass
        outputtext+="。"

        sheet['B12']=outputtext
        tonow = datetime.datetime.now()
        sheet['D3']=str(tonow.month) + "/" + str(tonow.day)


        wb1.save(targetpath)

        logger.warning('報告保存完畢')
        try:
            os.remove(testlistpath)
            logger.warning('刪除多餘檔案:測試用例(xlsx)')
        except OSError as e:
            print(e)
        try:
            os.remove(downloadpath+"/tempfile123.html")
            logger.warning('刪除多餘檔案:測試用例(html)')
        except OSError as e:
            print(e)
        except:
            pass
        logger.info("報告生成完畢")
    except:
        logger.error("發生錯誤",exc_info=True)
    logger.warning("報告生成完畢")


def testcase_getter(url,account,password,downloadpath):
    try:
        os.remove(downloadpath+"/tempfile123.xlsx")
        print("刪除舊xlsx檔成功")
        logger.warning('刪除舊檔案:測試用例(xlsx)')
    except:
        pass
    try:
        os.remove(downloadpath+"/tempfile123.html")
        print("刪除舊html檔成功")
        logger.warning('刪除舊檔案:測試用例(html)')
    except:
        pass

    if url!="":
        a=downloadpath.replace("/","\\")
        loginurl='http://192.168.21.239/zentao/user-login.html'
        session_requests = requests.session()
        session_requests.get(loginurl)
        SALT = '2111727669'
        headers = {
            }
        payload = {
                'account': account,
                'password':password,
                'referer': '',
                'verifyRand': SALT,
                'keepLogin': 1
            } 
        response = session_requests.request("POST", loginurl, headers=headers, data=payload)

        id_list = url[url.find('-cases-')+7:url.find('.html')]
        response = session_requests.request("GET", url)
        case_num = str(response.cookies['preTaskID'])
        payload = {
            'fileName': 'tempfile123',
            'fileType': 'html',
            'exportType': 'all',
            'exportFields[]': ['lastRunResult','id','module','title','real','lastRunner']
        }
        response = session_requests.request("POST", f'http://192.168.21.239/zentao/testcase-export-{case_num}-%60case%60_desc-{id_list}.html',data=payload)
        with open(os.path.join(downloadpath,'tempfile123.html'),'wb') as file:
            file.write(response.content)
        logger.warning('下載完成')
        return response.text
    else:
        logger.warning('無禪道測試單')


def parse_args():
    parser = ArgumentParser(prog='main.py') 
    parser.add_argument('--testcase_url', '-caseurl', default='', type=str, required=False, help='testcase url')
    parser.add_argument('--jira_url', '-jurl', default='', type=str, required=False, help='jira url')
    return parser.parse_args()



if __name__ == '__main__':
    args = parse_args() #從外部取值
    testcaseurl=args.testcase_url
    jiraurl=args.jira_url
    path='./Report' #寫死，是給 Jenkins 下載 Report 的地方
    testcase_getter(testcaseurl,'oliverchiu','!QAZ2wsx',path)
    generate_report('./Report',"./Report/result.xlsx",jiragetbug.get_jirabug(account='oliver206',password='XD6R247L',urls=jiraurl,bugstatus=3),jiraurl=jiraurl)